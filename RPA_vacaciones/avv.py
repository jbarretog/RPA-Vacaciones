import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

def ejecutar_automatizacion():
    # Ruta de los archivos
    archivo = 'InformeMaxtimeDosMeses.xlsx'  # Suponiendo que este archivo tiene datos de dos meses
    bdVacaciones = 'ReporteAplicativoVacaciones.xlsm'  # Ajustado a la extensión correcta
    output_file = 'practicas01_filtrado_con_fechas.xlsx'
    new_sheet = 'Filtrado_Falso'

    # Verificar existencia de archivos
    if not os.path.isfile(archivo):
        messagebox.showerror("Error", f"El archivo {archivo} no se encuentra en la ruta especificada.")
        return

    if not os.path.isfile(bdVacaciones):
        messagebox.showerror("Error", f"El archivo {bdVacaciones} no se encuentra en la ruta especificada.")
        return

    # Leer los archivos Excel
    try:
        dfMaxTime = pd.read_excel(archivo, skiprows=4)
        dfVacaciones = pd.read_excel(bdVacaciones, sheet_name='Hoja1')  # Ajustar el nombre de la hoja 
    except Exception as e:
        messagebox.showerror("Error", f"Error al leer los archivos: {e}")
        return

    # Filtrar datos para los meses relevantes
    filtro = dfMaxTime[(dfMaxTime['Actividad'] == 'NOV-VACACIONES')]

    # Ordenar dfVacaciones por Identificacion y fecha de inicio
    dfVacaciones = dfVacaciones.sort_values(['Identificacion', 'Fecha_inicio_vacaciones'], ascending=[True, False]).drop_duplicates('Identificacion')

    # Realizar el left join
    resultado = pd.merge(filtro, dfVacaciones[['Identificacion', 'Fecha_inicio_vacaciones', 'Fecha_fin_vacaciones']],
        left_on='Cedula', right_on='Identificacion', how='left')

    # Eliminar la columna duplicada de Identificacion si es necesario
    resultado = resultado.drop('Identificacion', axis=1)

    # Insertar una nueva columna de fecha de los campos separados
    resultado['Reporte_maxtime'] = pd.to_datetime({
        'year': resultado['Año'],
        'month': resultado['Mes'],
        'day': resultado['Dia']
    }).dt.strftime('%d/%m/%Y')  # Mantener el formato 'dd/mm/yyyy'

    # La nueva columna se llamará 'validación' y evaluará la condición lógica
    resultado['validacion'] = (pd.to_datetime(resultado['Reporte_maxtime'], format='%d/%m/%Y') >= resultado['Fecha_inicio_vacaciones']) & \
        (pd.to_datetime(resultado['Reporte_maxtime'], format='%d/%m/%Y') <= resultado['Fecha_fin_vacaciones'])

    # Convertir todas las columnas de tipo datetime al formato corto dd/mm/yyyy
    for col in resultado.columns:
        if pd.api.types.is_datetime64_any_dtype(resultado[col]):
            resultado[col] = resultado[col].dt.strftime('%d/%m/%Y')

    # Guardar el resultado
    resultado.to_excel(output_file, index=False)

    # Filtrar filas donde 'validacion' es False
    df_falso = resultado[resultado['validacion'] == False]

    # Convertir 'Reporte_maxtime' a datetime para filtrar por mes y año actuales
    df_falso['Reporte_maxtime'] = pd.to_datetime(df_falso['Reporte_maxtime'], format='%d/%m/%Y')

    # Obtener el mes actual
    mes_actual = datetime.now().month
    ano_actual = datetime.now().year

    # Filtrar df_falso para el mes actual
    df_falso_mes_actual = df_falso[(df_falso['Reporte_maxtime'].dt.month == mes_actual) & 
                                   (df_falso['Reporte_maxtime'].dt.year == ano_actual)]

    # Guardar las filas con validacion False y mes actual en una nueva hoja
    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_falso_mes_actual.to_excel(writer, sheet_name=new_sheet, index=False)

    # Aplicar formato en la nueva hoja
    wb = load_workbook(output_file)
    ws = wb[new_sheet]

    # Definir el color amarillo para el formato
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Pintar las celdas con 'False' en amarillo
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.min_column, max_col=ws.max_column):
        for cell in row:
            if cell.value == False:
                cell.fill = yellow_fill

    # Guardar el archivo con el formato aplicado
    wb.save(output_file)

    # Abrir el archivo automáticamente en Windows
    os.startfile(output_file)
    
    # Mensaje de éxito
    messagebox.showinfo("Éxito", "La automatización se ha completado exitosamente.")

# Crear la interfaz gráfica
root = tk.Tk()
root.title("Automatización de Reportes de Vacaciones")
root.geometry("400x200")

# Crear y colocar el botón
run_button = tk.Button(root, text="Ejecutar Automatización", command=ejecutar_automatizacion, height=2, width=30)
run_button.pack(pady=60)

# Iniciar el bucle principal de la interfaz
root.mainloop()
